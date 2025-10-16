import pandas as pd
import pickle
import sys
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# ───────────────────────────────────────────────
# CLI usage
# ───────────────────────────────────────────────
if len(sys.argv) != 3:
    print("Usage: python 02.py <data_file.pkl> <params_file.json>")
    sys.exit(1)

DATA_PKL = sys.argv[1]
PARAMS_JSON = sys.argv[2]

with open(DATA_PKL, "rb") as f:
    payload = pickle.load(f)

results = payload.get("results")
if not results:
    print("[ERROR] No pivot results found in pickle.")
    sys.exit(1)

measure_year = payload.get("measure_year", "")
run_date_str = datetime.today().strftime("%Y-%m-%d")
output_path = f"CCBHC_Time_to_Services_I_SERV_Quality_Measure_report_{run_date_str}_MY{measure_year}.xlsx"

# ───────────────────────────────────────────────
# Sheet-friendly titles
# ───────────────────────────────────────────────
def clean_sheet_name(text):
    return text.replace(":", "").replace("/", "-").strip()[:30]

# ───────────────────────────────────────────────
# Auto-fit column widths
# ───────────────────────────────────────────────
def autofit_column_widths(ws, buffer=2):
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_len = len(str(cell.value))
                col_letter = get_column_letter(cell.column)
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), val_len + buffer)
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

# ───────────────────────────────────────────────
# Write to Excel with openpyxl
# ───────────────────────────────────────────────
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

# Write pivot tables with descriptions
for measure_label, age_dict in results.items():
    sheet_name = clean_sheet_name(measure_label)
    ws = wb.create_sheet(title=sheet_name)

    # Add a human-readable description at the top
    if measure_label == "I-SERV-1":
        ws.append(["I-SERV-1: Time from intake appointment to first clinical assessment for new clients."])
        ws.append(["Lower averages and higher percentages indicate improved performance."])
    elif measure_label == "I-SERV-2":
        ws.append(["I-SERV-2: Time from intake appointment to first billable service for new clients."])
        ws.append(["Lower averages and higher percentages indicate improved performance."])
    ws.append([])

    ws.append([f"{measure_label} – 12–17 year olds"])
    df_12_17 = age_dict.get("12-17 yo")
    if df_12_17 is not None:
        for r in dataframe_to_rows(df_12_17, index=True, header=True):
            ws.append(r)

    ws.append([])

    ws.append([f"{measure_label} – 18+ year olds"])
    df_18_plus = age_dict.get("18+ yo")
    if df_18_plus is not None:
        for r in dataframe_to_rows(df_18_plus, index=True, header=True):
            ws.append(r)

    autofit_column_widths(ws)

# Write raw data sheets with date formatting
df_raw_eval = payload.get("raw_eval")
df_raw_service = payload.get("raw_service")

def format_dates(df):
    for col in ["Appointment Date", "Appointment Created Date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d")
    return df

if df_raw_eval is not None:
    df_raw_eval = format_dates(df_raw_eval.copy())
    ws_eval = wb.create_sheet(title="I-SERV-1 Raw")
    for r in dataframe_to_rows(df_raw_eval, index=False, header=True):
        ws_eval.append(r)
    autofit_column_widths(ws_eval)

if df_raw_service is not None:
    df_raw_service = format_dates(df_raw_service.copy())
    ws_service = wb.create_sheet(title="I-SERV-2 Raw")
    for r in dataframe_to_rows(df_raw_service, index=False, header=True):
        ws_service.append(r)
    autofit_column_widths(ws_service)

wb.active = 0
wb.save(output_path)
print(f"[OK] Pivot tables and raw data written to {output_path}")
