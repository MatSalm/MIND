import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2

# Load the sampled DataFrame from temp_data.pkl
sampled_df = pd.read_pickle('temp_data.pkl')

# Define a dictionary with current column names as keys and new column names as values
column_renames = {
    'PATID': 'Client ID',
    'program_value': 'Program'
}

# Rename the columns
sampled_df.rename(columns=column_renames, inplace=True)

# Create a new workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Random Sample Report"

# Remove gridlines
ws.sheet_view.showGridLines = False

# Insert the logo image into cell A1
logo_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large.png'
logo = Image(logo_path)
logo.width = int(logo.width * 0.16)
logo.height = int(logo.height * 0.16)
ws.add_image(logo, 'A1')

# Insert header text in merged cells E4 to I4
filter_date = datetime.now().strftime('%Y-%m-%d')
header_text = f'Random Sample Report for {filter_date}'
ws.merge_cells('E4:I4')
header_cell = ws['E4']
header_cell.value = header_text
header_cell.alignment = Alignment(horizontal='center', vertical='center')
header_cell.font = Font(bold=True, size=16)

# Write the DataFrame to the sheet starting at cell D8
start_row = 8
start_col = 4
for r_idx, row in enumerate(dataframe_to_rows(sampled_df, index=False, header=True), start=start_row):
    for c_idx, value in enumerate(row, start=start_col):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.alignment = Alignment(horizontal='left')
        
        # Make header row bold
        if r_idx == start_row:
            cell.font = Font(bold=True)
        
        # Apply date formatting
        if cell.column_letter == 'E' and r_idx > start_row:
            cell.number_format = FORMAT_DATE_YYYYMMDD2
        
        # Apply number formatting to 'Client ID'
        if cell.column_letter == 'D' and r_idx > start_row:
            cell.number_format = '0'

# Determine the max length for each column
col_lengths = {}
for col in range(start_col, start_col + len(sampled_df.columns)):
    max_length = 0
    for row in range(start_row, start_row + len(sampled_df) + 1):  # Include headers
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    col_lengths[col] = max_length + 2  # Add extra spaces

# Auto-adjust column width
for col, length in col_lengths.items():
    ws.column_dimensions[ws.cell(row=start_row, column=col).column_letter].width = length * 1.2

# Add a border around the data excluding the headers
first_row = start_row + 1
last_row = start_row + len(sampled_df)
first_col = start_col
last_col = start_col + len(sampled_df.columns) - 1

for row in range(first_row, last_row + 1):
    ws.cell(row=row, column=first_col).border = Border(left=Side(style='thin'))
    ws.cell(row=row, column=last_col).border = Border(right=Side(style='thin'))
for col in range(first_col, last_col + 1):
    ws.cell(row=first_row, column=col).border = Border(top=Side(style='thin'))
    ws.cell(row=last_row, column=col).border = Border(bottom=Side(style='thin'))

# Ensure corners have the correct borders
ws.cell(row=first_row, column=first_col).border = Border(left=Side(style='thin'), top=Side(style='thin'))
ws.cell(row=first_row, column=last_col).border = Border(right=Side(style='thin'), top=Side(style='thin'))
ws.cell(row=last_row, column=first_col).border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
ws.cell(row=last_row, column=last_col).border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

# Save the workbook to a file with the specified filename
output_path = f'All_Program_Regulatory_Audit_random_sample_{filter_date}.xlsx'
wb.save(output_path)

print(f"Random sample Excel file created: {output_path}")
