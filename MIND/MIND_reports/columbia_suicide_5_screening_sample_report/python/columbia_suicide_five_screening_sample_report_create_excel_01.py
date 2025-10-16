import pandas as pd
import os
import sys
import pickle
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import datetime

# Load the sampled DataFrame from temp_data.pkl
try:
    sampled_df = pd.read_pickle('temp_data.pkl')
except Exception as e:
    print(f"Error loading sample data from temp_data.pkl: {e}")
    sys.exit(1)

# Rename the column to "Client ID"
sampled_df.rename(columns={'PATID': 'Client ID'}, inplace=True)

# Load date parameters from temp_params.json
params_file_path = 'temp_params.json'  # Ensure the path matches where the file is stored
try:
    with open(params_file_path, 'r') as file:
        params = json.load(file)
    start_date = datetime.strptime(params['start_date'], '%Y-%m-%d')
    end_date = datetime.strptime(params['end_date'], '%Y-%m-%d')
except Exception as e:
    print(f"Error reading date parameters from JSON: {e}")
    sys.exit(1)

# Create a new workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Random Sample Report"

# Remove gridlines
ws.sheet_view.showGridLines = False

# Insert the logo image into cell A1
logo_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large.png'
logo = Image(logo_path)
logo.width, logo.height = logo.width * 0.16, logo.height * 0.16
ws.add_image(logo, 'A1')

# Insert header text in merged cells and auto-adjust to fit the text
header_text = f'Columbia Screening Sample {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
ws.merge_cells('E4:L4')
header_cell = ws['E4']
header_cell.value = header_text
header_cell.alignment = Alignment(horizontal='center', vertical='center')
header_cell.font = Font(bold=True, size=16)

# Write only the "Client ID" column to the sheet starting at cell H8
start_row, start_col = 8, 8
for r_idx, row in enumerate(dataframe_to_rows(sampled_df[['Client ID']], index=False, header=True), start=start_row):
    for c_idx, value in enumerate(row, start=start_col):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == start_row:  # Header row: Only bottom border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))
        else:
            cell.alignment = Alignment(horizontal='left')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
            if r_idx == start_row + len(sampled_df):  # Bottom border for the last row
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
            if c_idx == start_col:  # Format "Client ID" as number
                cell.number_format = '0'

# Ensure the bottom row has left, right, and bottom borders
for c_idx in range(start_col, start_col + 1):
    cell = ws.cell(row=start_row + len(sampled_df), column=c_idx)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

# Determine the max length for each column and auto-adjust column width
for col in ws.iter_cols(min_row=start_row, max_row=start_row + len(sampled_df), min_col=start_col, max_col=start_col):
    max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2  # Add extra spaces for padding

# Save the workbook using the date range in the filename
output_filename = f'columbia_screening_random_sample_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx'
try:
    wb.save(output_filename)
    print(f"Excel file saved as {output_filename}")
except Exception as e:
    print(f"Error saving Excel file: {e}")
    sys.exit(1)
