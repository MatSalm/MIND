import os
import sys
import pandas as pd
import pickle
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
import configparser
from dotenv import load_dotenv
from pathlib import Path
from PIL import Image

# ------------------- Additional imports for advanced Excel formatting -------------------
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2
# ---------------------------------------------------------------------------------------

# Load environment variables from the .env file
load_dotenv()

# Ensure proper usage by checking the number of command-line arguments
if len(sys.argv) != 3:
    print("Usage: python 01.py <data_file> <param_file>")
    sys.exit(1)

# Extract the data and parameter file paths from the command-line arguments
data_file = sys.argv[1]
param_file = sys.argv[2]

try:
    # Load the DataFrames from the .pkl file
    if os.path.exists(data_file):
        with open(data_file, 'rb') as f:
            data = pickle.load(f)
            df_all_clients = data.get('all_clients')
            df_clients_to_discharge = data.get('clients_to_discharge')
    else:
        raise FileNotFoundError(f"{data_file} not found.")

    # Debug: Check if DataFrames are None or empty
    if df_all_clients is None:
        print("Error: 'all_clients' DataFrame is None.")
        sys.exit(1)
    else:
        print(f"All Clients DataFrame loaded successfully with {len(df_all_clients)} records.")

    if df_clients_to_discharge is None:
        print("Error: 'clients_to_discharge' DataFrame is None.")
        sys.exit(1)
    else:
        print(f"Clients to Discharge DataFrame loaded successfully with {len(df_clients_to_discharge)} records.")

    # Convert 'Last Date of Service' to date format
    df_all_clients['Last Date of Service'] = pd.to_datetime(df_all_clients['Last Date of Service']).dt.date
    df_clients_to_discharge['Last Date of Service'] = pd.to_datetime(df_clients_to_discharge['Last Date of Service']).dt.date

    # Read parameters from the JSON parameter file
    with open(param_file, 'r', encoding='utf-8') as f:
        parameters = json.load(f)

    # Read email configuration from config.ini
    config_file_path = os.path.join('..', 'config', 'config.ini')
    config = configparser.ConfigParser()
    config.read(config_file_path)
    to_email = config.get('email', 'to_email')
    days_since_last_service = config.get('report', 'days_since_last_service')

    # Get report date from parameters
    report_date = parameters.get('report_date', 'Unknown Date')

    # Prepare the Excel filename
    excel_file_name = f"CCBHC_Client_Last_Seen_Date_Potentially_to_Discharge_Report_on_{report_date}.xlsx"

    # -------------------- Define border styles --------------------
    thin_left_border = Border(left=Side(style='thin'))
    thin_right_border = Border(right=Side(style='thin'))
    thin_top_border = Border(top=Side(style='thin'))
    thin_bottom_border = Border(bottom=Side(style='thin'))
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # -------------------------------------------------------------

    # -------------------- Create a new workbook --------------------
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)  # Remove the default sheet

    # Path to your logo (if needed in the Excel file)
    logo_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large.png'

    def create_formatted_sheet(workbook, df, sheet_title):
        """
        Creates a new sheet with the provided DataFrame, applying the style from your example script.
        """
        sheet = workbook.create_sheet(title=sheet_title)

        # Remove gridlines
        sheet.sheet_view.showGridLines = False

        # Insert the logo image in A1 (if file exists)
        if os.path.exists(logo_path):
            logo = ExcelImage(logo_path)
            # Scale down the image
            logo.width = int(logo.width * 0.16)
            logo.height = int(logo.height * 0.16)
            sheet.add_image(logo, 'A1')

        # Merge cells E4 to I4 for the title
        sheet.merge_cells('E4:I4')
        header_cell = sheet['E4']
        # Example: "CCBHC Clients to Discharge" or "All Clients" plus date
        header_cell.value = f"{sheet_title} Report for {report_date}"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = Font(bold=True, size=16)

        # Start writing the DataFrame at row 8, column 4
        start_row = 8
        start_col = 4

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='left')

                # Bold headers
                if r_idx == start_row:
                    cell.font = Font(bold=True)

                # If this is the 'Last Date of Service' column (which is column E if start_col=4 and the date col is E):
                # We can do a more general check if needed:
                if cell.column_letter == 'E' and r_idx > start_row:
                    # Date formatting
                    cell.number_format = FORMAT_DATE_YYYYMMDD2

        # Auto-adjust column width
        num_rows = len(df) + 1  # plus headers
        num_cols = len(df.columns)
        for col in range(start_col, start_col + num_cols):
            max_length = 0
            for row in range(start_row, start_row + num_rows):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            sheet.column_dimensions[sheet.cell(row=start_row, column=col).column_letter].width = max_length + 2

        # Add a border around the data excluding the headers
        first_row = start_row + 1
        last_row = start_row + len(df)
        first_col = start_col
        last_col = start_col + num_cols - 1

        # Left & right borders
        for row in range(first_row, last_row + 1):
            sheet.cell(row=row, column=first_col).border = Border(left=thin_border.left)
            sheet.cell(row=row, column=last_col).border = Border(right=thin_border.right)

        # Top & bottom borders
        for col in range(first_col, last_col + 1):
            sheet.cell(row=first_row, column=col).border = Border(top=thin_border.top)
            sheet.cell(row=last_row, column=col).border = Border(bottom=thin_border.bottom)

        # Ensure corners have the correct borders
        sheet.cell(row=first_row, column=first_col).border = Border(left=thin_border.left, top=thin_border.top)
        sheet.cell(row=first_row, column=last_col).border = Border(right=thin_border.right, top=thin_border.top)
        sheet.cell(row=last_row, column=first_col).border = Border(left=thin_border.left, bottom=thin_border.bottom)
        sheet.cell(row=last_row, column=last_col).border = Border(right=thin_border.right, bottom=thin_border.bottom)

        return sheet

    # Create the two sheets with your DataFrames
    create_formatted_sheet(wb, df_clients_to_discharge, "CCBHC Program Client Discharge")
    create_formatted_sheet(wb, df_all_clients, "All Clients")

    # Save the workbook
    wb.save(excel_file_name)

    # ------------------- Email settings & signature remain unchanged -------------------

    smtp_email = os.getenv('EMAIL_smtp_email')
    smtp_server = os.getenv('EMAIL_smtp_server')
    smtp_port = os.getenv('EMAIL_smtp_port')

    # Check if environment variables are loaded correctly
    if not smtp_email or not smtp_port or not smtp_server:
        raise ValueError("SMTP configuration is missing in the environment variables.")

    smtp_port = int(smtp_port)

    # Create the email subject and body
    subject = f"[secure] CCBHC Non-Discharged Clients to Potentially Be Discharged Report - Report Date: {report_date}"
    body = f"""
    <html>
      <body style="font-family: 'Segoe UI', sans-serif; color: #242424;">
        <p>Hello,</p>
        <p style="margin-left: 40px;">Attached is the report of non-discharged clients who may or may not have a NOMs baseline completed and have not been seen in {days_since_last_service} days and may be recommended for discharge as of {report_date}.</p>
        <p>Thank you,</p>
        <div style="margin-top: 20px;">
          <table cellspacing="0" cellpadding="0" border="0" style="width: auto;">
            <tr>
              <td colspan="2">
                <hr style="border: none; border-top: 1px solid #e0e0e0; margin: 10px 0;">
              </td>
            </tr>
            <tr>
              <td style="vertical-align: top; padding-right: 10px;">
                <p style="margin: 0; font-size: 14px;"><strong>MIND</strong><br>
                Enterprise Reporting System,<br>
                Hillcrest Family Services</p>
                <p style="margin: 5px 0;">
                  <a href="https://hillcrest-fs.org/" style="color: #467886; text-decoration: none;">hillcrest-fs.org</a>
                </p>
              </td>
              <td style="border-left: 1px solid #e0e0e0; padding-left: 10px; vertical-align: top;">
                <img src="cid:image1" style="height: 50px; margin-top: 10px;">
                <img src="cid:image2" style="height: 50px; margin-left: 10px; margin-top: 10px;">
              </td>
            </tr>
          </table>
        </div>
      </body>
    </html>
    """

    # Create the email message
    msg = MIMEMultipart('related')
    msg['From'] = smtp_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    # Attach the Excel file
    with open(excel_file_name, 'rb') as attachment:
        part = MIMEApplication(attachment.read(), Name=os.path.basename(excel_file_name))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(excel_file_name)}"'
        msg.attach(part)

    # Function to resize images
    def resize_image(image_path, output_path, size):
        with Image.open(image_path) as img:
            img.thumbnail(size)
            img.save(output_path, format='PNG')

    # Paths to the images (unchanged)
    signature_image_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large.png'
    resized_signature_image_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large_resized.png'
    additional_image_path = r'C:\MIND\MIND\MIND_images\MIND_logo.png'
    resized_additional_image_path = r'C:\MIND\MIND\MIND_images\MIND_logo_resized.png'

    # Resize and embed the signature image in the email
    resize_image(signature_image_path, resized_signature_image_path, (200, 200))
    with open(resized_signature_image_path, 'rb') as img:
        msg_image = MIMEImage(img.read())
        msg_image.add_header('Content-ID', '<image1>')
        msg.attach(msg_image)

    # Resize and embed the additional image in the email
    resize_image(additional_image_path, resized_additional_image_path, (130, 130))
    with open(resized_additional_image_path, 'rb') as img:
        msg_image2 = MIMEImage(img.read())
        msg_image2.add_header('Content-ID', '<image2>')
        msg.attach(msg_image2)

    # Set up the SMTP server and send the email
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.ehlo()
    # Authentication is removed as it’s not required
    server.send_message(msg)
    server.quit()

    print(f"Email sent successfully to {to_email}.")

    # Cleanup: Delete the Excel file and resized images
    os.remove(excel_file_name)
    os.remove(resized_signature_image_path)
    os.remove(resized_additional_image_path)
    os.remove(data_file)
    os.remove(param_file)

except Exception as e:
    # Print and raise any exceptions encountered during the script execution
    print(f"Error: {str(e)}")
    raise
