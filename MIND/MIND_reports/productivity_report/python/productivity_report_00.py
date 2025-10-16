import os
import re
import sys
import glob
import pyodbc
import configparser
import pandas as pd
import numpy as np
from email import encoders
import shutil
from datetime import datetime, timedelta, date
import datetime as dt
from dotenv import load_dotenv
from pathlib import Path
import pickle
import json

# Load environment variables from MIND.env file
load_dotenv(dotenv_path='C:/MIND/MIND/MIND_config/MIND.env')

# Function to validate the config files for security check
def is_valid_file_path(file_path):
    """Checks if the provided file path is valid and safe."""
    if not os.path.isabs(file_path):
        return False

    # Basic URL check with regex
    url_regex = r"^https?://[^\s/\\]+\.[^\s/\\]+$"  # Basic https?://domain.extension format
    if re.match(url_regex, file_path):
        return False

    return os.path.exists(file_path)  # Check if file exists

# Read parameters passed from MIND.py
if len(sys.argv) < 3:
    raise ValueError("This script requires at least 2 arguments: data file path and parameter file path.")

data_file_path = sys.argv[1]
param_file_path = sys.argv[2]

# Load data from the pickle file if it exists
if os.path.exists(data_file_path):
    with open(data_file_path, 'rb') as f:
        data = pickle.load(f)
else:
    data = None

# Load parameters from the JSON file
with open(param_file_path, 'r') as f:
    parameters = json.load(f)

# Get the report configuration file path from parameters
report_config_file_path = parameters.get("report_config_file_path")
if not is_valid_file_path(report_config_file_path):
    raise ValueError("Invalid report config.ini file path!")

# Read report-specific configuration
report_config = configparser.ConfigParser(interpolation=None)
report_config.read(report_config_file_path)

# Get report-specific parameters
past_days = report_config["calendar"]["past_days"]
# Ensure the values are always treated as raw strings
calendar_start_date = report_config.get("calendar", "calendar_start_date", fallback="").strip()
calendar_stop_date = report_config.get("calendar", "calendar_stop_date", fallback="").strip()

appointment_status = report_config["report"]["appointment_status"]
available_time_sites = report_config["report"]["available_time_sites"]
productivity_service_code_list_location = report_config["report"]["productivity_service_code_list_location"]
productivity_payout_percentage_list_location = report_config["report"]["productivity_payout_percentage_list_location"]
user_roles = report_config["report"]["user_roles"]
exception_service_codes = report_config["report"]["exception_service_codes"]

# Get the report calendar parameters for calculating Working Hours
therapist_availability = report_config["working_hours"]["therapist_availability"]
prescriber_availability = report_config["working_hours"]["prescriber_availability"]
staff_expected_availability = report_config["report"]["staff_expected_availability"]
staff_expected_availability = pd.read_csv(staff_expected_availability, dtype={'staff': str})
staff_working_days_hours = report_config["report"]["staff_working_days_hours"]
staff_working_days_hours = pd.read_csv(staff_working_days_hours)

# Get the productive time parameters for calculating productive hours
draft_final_code = report_config["productive_time"]["draft_final_code"]
document_routing_status = report_config["productive_time"]["document_routing_status"]
use_date_of_service = report_config["productive_time"]["use_date_of_service"]
use_date_of_note = report_config["productive_time"]["use_date_of_note"]
use_note_table = report_config["productive_time"]["use_note_table"]
use_note_billing_charge_table = report_config["productive_time"]["use_note_billing_charge_table"]

# Get database credentials from environment variables
server = os.getenv("database_server")
port = os.getenv("database_port")
databasePM = os.getenv("databasePM")
username = os.getenv("database_username")
password = os.getenv("database_password")

# Get the Email credentials from environment variables
smtp_server = os.getenv("EMAIL_smtp_server")
smtp_port = os.getenv("EMAIL_smtp_port")
smtp_password = os.getenv("EMAIL_smtp_password")
smtp_email = os.getenv("EMAIL_smtp_email")
to_email = report_config.get('email', 'to_email')

conn = None
df = None  # Define df here

# Construct the connection string
conn_stringPM = f'DRIVER={{InterSystems IRIS ODBC35}};SERVER={server};PORT={port};DATABASE={databasePM};UID={username};PWD={password}'

# -------------------------
# Fixing date handling logic
# -------------------------

def is_valid_date(date_input):
    """Check if the input is a valid date, datetime object, or a valid date string."""
    if date_input is None or (isinstance(date_input, str) and not date_input.strip()):
        return False  # Empty or None input
    if isinstance(date_input, (dt.datetime, dt.date)):  
        return True  # Already a valid date/datetime object
    if isinstance(date_input, str):
        try:
            dt.datetime.strptime(date_input, '%Y-%m-%d')  # Validate the date format
            return True
        except ValueError:
            return False  # Invalid date format
    return False  # If it's neither a string nor a date/datetime object




# Ensure the values are strings before parsing
calendar_start_date = str(calendar_start_date) if isinstance(calendar_start_date, (str, date)) else None
calendar_stop_date = str(calendar_stop_date) if isinstance(calendar_stop_date, (str, date)) else None

# Debugging output to verify types
print(f"DEBUG: calendar_start_date type -> {type(calendar_start_date)}, value -> {calendar_start_date}")
print(f"DEBUG: calendar_stop_date type -> {type(calendar_stop_date)}, value -> {calendar_stop_date}")

# Convert them to datetime objects if they are valid strings
if calendar_start_date and is_valid_date(calendar_start_date):
    calendar_start_date = datetime.strptime(calendar_start_date, '%Y-%m-%d').date()
else:
    calendar_start_date = None

if calendar_stop_date and is_valid_date(calendar_stop_date):
    calendar_stop_date = datetime.strptime(calendar_stop_date, '%Y-%m-%d').date()
else:
    calendar_stop_date = None

# Debugging output to verify conversion
print(f"DEBUG: calendar_start_date converted -> {type(calendar_start_date)}, value -> {calendar_start_date}")
print(f"DEBUG: calendar_stop_date converted -> {type(calendar_stop_date)}, value -> {calendar_stop_date}")
# Get today's date
today = datetime.now()


# Determine the date range based on today's date or calendar dates
if calendar_start_date and calendar_stop_date:
    start_date = calendar_start_date
    end_date = calendar_stop_date
elif 1 <= today.day <= 15:
    start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=16).date()
    end_date = (today.replace(day=1) - timedelta(days=1)).date()
else:
    start_date = today.replace(day=1).date()
    end_date = today.replace(day=15).date()

schedule_start_date = start_date
schedule_end_date = end_date









# Function to convert time string to datetime object
def convert_time_to_datetime(time_str):
    return dt.datetime.strptime(time_str, "%I:%M %p")

def convert_to_military_time(time_str):
    if time_str is None:
        return None
    time_str = re.sub(r'(?<=\d)(?=[APap][Mm])', ' ', time_str)
    time_str = time_str.upper().strip()
    datetime_obj = dt.datetime.strptime(time_str, "%I:%M %p")
    return datetime_obj.strftime("%H:%M")

# Convert available_time_sites to a list of sites and remove leading/trailing whitespaces
available_time_sites = [site.strip() for site in available_time_sites.split(',')]

# Dictionary to store available hours (key: practitioner ID, value: list of dictionaries)
practitioner_availability = {}

# Mapping of day_code to match Python's weekday() function
day_code_mapping = {
    1: 6,  # Sunday to 6
    2: 0,  # Monday to 0
    3: 1,  # Tuesday to 1
    4: 2,  # Wednesday to 2
    5: 3,  # Thursday to 3
    6: 4,  # Friday to 4
    7: 5   # Saturday to 5
}


# # Check if calendar_start_date and calendar_stop_date are valid dates
# start_date = dt.datetime.strptime(calendar_start_date, "%Y-%m-%d").date()
# end_date = dt.datetime.strptime(calendar_stop_date, "%Y-%m-%d").date()

# Process each row of the staff_working_days_hours DataFrame
for index, row in staff_working_days_hours.iterrows():
    practitioner_id = row['STAFFID']
    day_code = day_code_mapping[int(row['day_code'])]
    site_name = row['site_name']
    start_time_military = convert_to_military_time(row['staff_start_time'])
    end_time_military = convert_to_military_time(row['staff_end_time'])
    staff_name = row['staff_name']
    unique_ID = row['ID']
    user_role = row['USERROLE']

    # Only process this row if site_name is in available_time_sites
    if site_name in available_time_sites:
        # Create list for practitioner if it doesn't exist
        if practitioner_id not in practitioner_availability:
            practitioner_availability[practitioner_id] = []

        # Loop through the date range and check for matching day code
        date = start_date
        while date <= end_date:
            if date.weekday() == day_code:
                # Append a dictionary instead of a list
                practitioner_availability[practitioner_id].append({
                    'Practitioner ID': practitioner_id,
                    'date': date,
                    'start_time': start_time_military,
                    'end_time': end_time_military,
                    'staff_name': staff_name,
                    'site_name': site_name,
                    'unique_ID': unique_ID,
                    'User Role': user_role
                })
            date += dt.timedelta(days=1)

# Create a list to store the availability data
data = []

for practitioner_id, availability in practitioner_availability.items():
    if availability:  # Check if availability list is not empty
        for entry in availability:
            data.append({
                'Practitioner ID': practitioner_id,
                'Date': entry['date'],
                'Start Time': entry['start_time'],
                'End Time': entry['end_time'],
                'Staff Name': entry['staff_name'],
                'Site': entry['site_name'],
                'ID': entry['unique_ID'],
                'User Role': entry['User Role']
            })

# Convert the list into a DataFrame
availability_df = pd.DataFrame(data)

# Explicitly set the data type for the 'Practitioner ID' column to object
availability_df['Practitioner ID'] = availability_df['Practitioner ID'].astype('object')

# # Function to check if a date is in 'yyyy-mm-dd' format
# def is_valid_date(date_str):
#     try:
#         datetime.strptime(date_str, '%Y-%m-%d')
#         return True
#     except ValueError:
#         return False

# Convert calendar_start_date and calendar_stop_date to datetime objects only if they are strings
if isinstance(calendar_start_date, str) and is_valid_date(calendar_start_date):
    calendar_start_date = datetime.strptime(calendar_start_date, '%Y-%m-%d').date()

if isinstance(calendar_stop_date, str) and is_valid_date(calendar_stop_date):
    calendar_stop_date = datetime.strptime(calendar_stop_date, '%Y-%m-%d').date()


# Determine the date range based on today's date or calendar dates
if calendar_start_date and calendar_stop_date:
    start_date = calendar_start_date
    end_date = calendar_stop_date
else:
    today = datetime.now()
    if 1 <= today.day <= 15:
        start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=16).date()
        end_date = (today.replace(day=1) - timedelta(days=1)).date()
    else:
        start_date = today.replace(day=1).date()
        end_date = today.replace(day=15).date()

schedule_start_date = start_date
schedule_end_date = end_date

# Extract the years involved in the reporting range
start_year = start_date.year
end_year = end_date.year

# List of fixed holiday dates (format: 'MM-DD')
holidays = ['01-01', '12-24', '12-25', '12-31']

# Function to calculate Easter Sunday
def calculate_easter(year):
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day)

# Function to calculate dynamic holidays
def get_holiday_dates(start_year, end_year):
    holiday_dates = []

    # Handle fixed holidays for both years
    for year in range(start_year, end_year + 1):
        for holiday in holidays:
            holiday_dates.append(datetime.strptime(f"{year}-{holiday}", "%Y-%m-%d"))

    # Dynamic holidays for each year
    for year in range(start_year, end_year + 1):
        # Good Friday (2 days before Easter)
        easter = calculate_easter(year)
        good_friday = easter - timedelta(days=2)
        holiday_dates.append(good_friday)

        # Memorial Day (Last Monday in May)
        last_monday_in_may = max(pd.date_range(start=f'{year}-05-01', end=f'{year}-05-31', freq='W-MON'))
        holiday_dates.append(last_monday_in_may)

        # Add Independence Day, Labor Day, Thanksgiving, Day after Thanksgiving
        thanksgiving = max(pd.date_range(start=f'{year}-11-01', end=f'{year}-11-30', freq='W-THU'))
        holiday_dates.extend([
            pd.Timestamp(f'{year}-07-04'),
            pd.Timestamp(f'{year}-09-01') + pd.offsets.Week(weekday=0),
            thanksgiving,
            thanksgiving + timedelta(days=1)
        ])

    return holiday_dates

# Get holidays for the reporting years
holiday_dates = get_holiday_dates(start_year, end_year)

# Adjust for holidays on weekends
adjusted_holiday_dates = set(holiday_dates)
for holiday in holiday_dates:
    if holiday.weekday() == 5:  # Saturday
        adjusted_holiday_dates.add(holiday - timedelta(days=1))  # Friday
    elif holiday.weekday() == 6:  # Sunday
        adjusted_holiday_dates.add(holiday + timedelta(days=1))  # Monday

# Filter holidays within the reporting range
adjusted_holiday_dates = [date for date in adjusted_holiday_dates if start_date <= date.date() <= end_date]

# Remove records with holidays
availability_df['Date'] = pd.to_datetime(availability_df['Date']).dt.date
original_count = len(availability_df)
availability_df = availability_df[~availability_df['Date'].isin([holiday.date() for holiday in adjusted_holiday_dates])]
removed_count = original_count - len(availability_df)

# Display the number of removed records
print(f"Number of records affected by holiday changes: {removed_count}")



#FIND THE EXCEPTION HOURS
# Create dataframe with list of EXCEPTION appointments as listed in the exception_service_codes parameter

# Establish database connection
conn = pyodbc.connect(conn_stringPM)

# Convert appointment_status to a list and remove leading/trailing whitespaces
appointment_status = [status.strip() for status in appointment_status.split(',')]

# Ensure exception_service_codes is a list
if isinstance(exception_service_codes, str):
    exception_service_codes = [code.strip() for code in exception_service_codes.split(',')]

# Check if exception_service_codes is not empty
if exception_service_codes:
    # Create the parameter placeholders
    placeholders = ', '.join('?' for _ in exception_service_codes)
    # Modify the SQL query to include SERVICE_CODE condition
    sql = f"""
    SELECT SERVICE_CODE, STAFFID, appointment_date, appointment_start_time, appointment_end_time, duration_minutes, location_value, location_code, program_value, service_description, site_name, recurring_indicator, status_value
    FROM AVPM.SYSTEM.AppointmentData
    WHERE appointment_date >= ? AND appointment_date <= ? AND SERVICE_CODE IN ({placeholders})
    """
    # Add start_date and end_date to the beginning of params list
    params = [start_date, end_date] + exception_service_codes
else:
    # Print message if no exception service codes are provided
    print("No exception service codes provided in config file.")
    # Modify the SQL query to exclude SERVICE_CODE condition
    sql = f"""
    SELECT SERVICE_CODE, STAFFID, appointment_date, appointment_start_time, appointment_end_time, duration_minutes, location_value, location_code, program_value, service_description, site_name, recurring_indicator, status_value
    FROM AVPM.SYSTEM.AppointmentData
    WHERE appointment_date >= ? AND appointment_date <= ?
    """
    params = [start_date, end_date]

# Execute the query and create the dataframe
exception_appointment_df = pd.read_sql(sql, conn, params=params)

# Close the database connection
conn.close()






# Function to convert to military time
def convert_to_military_time(time_str):
    if isinstance(time_str, str):
        time_str = re.sub(r'(?<=\d)(?=[APap][Mm])', ' ', time_str)
        time_str = time_str.upper().strip()
        datetime_obj = dt.datetime.strptime(time_str, "%I:%M %p")
        return datetime_obj.strftime("%H:%M")
    return time_str

# Create a copy of availability_df
availability_copy_df = availability_df.copy()

# Pad Practitioner ID with leading zeros to match STAFFID format
availability_copy_df['Practitioner ID'] = availability_copy_df['Practitioner ID'].astype(str).str.zfill(6)

# Convert date columns to datetime
availability_copy_df['Date'] = pd.to_datetime(availability_copy_df['Date'])

exception_appointment_df['appointment_date'] = pd.to_datetime(exception_appointment_df['appointment_date'])
exception_appointment_df['appointment_start_time'] = exception_appointment_df['appointment_start_time'].apply(convert_to_military_time)
exception_appointment_df['appointment_end_time'] = exception_appointment_df['appointment_end_time'].apply(convert_to_military_time)

# Function to filter appointments
def filter_appointments(exception_df, availability_df):
    # Create a set of valid (Practitioner ID, Date) pairs from the availability dataframe
    valid_pairs = set(zip(availability_df['Practitioner ID'], availability_df['Date']))
    
    # Filter the exception dataframe to only include rows with valid (STAFFID, appointment_date) pairs
    exception_df = exception_df[exception_df.apply(lambda row: (row['STAFFID'], row['appointment_date']) in valid_pairs, axis=1)]
    
    return exception_df

exception_appointment_df = exception_appointment_df[
    exception_appointment_df['site_name'].isin(available_time_sites)
]

# Apply the filter function to exception_appointment_df
exception_appointment_df = filter_appointments(exception_appointment_df, availability_copy_df)



# print("JESSICA rows by site:")
# print(
#     exception_appointment_df[
#         exception_appointment_df['STAFFID'] == '000757'
#     ][['site_name', 'duration_minutes']]
#     .groupby('site_name')
#     .sum()
# )






# Define function to calculate availability using NetSmart available working hours
def calculate_working_hours(availability_df):
    # Convert 'Start Time' and 'End Time' to datetime format
    availability_df['Start Time'] = pd.to_datetime(availability_df['Start Time'], format='%H:%M', errors='coerce').dt.time
    availability_df['End Time'] = pd.to_datetime(availability_df['End Time'], format='%H:%M', errors='coerce').dt.time

    # Calculate 'Working Hours' with handling for NaN values
    def calculate_hours(row):
        if pd.isna(row['Start Time']) or pd.isna(row['End Time']):
            return 0  # or any appropriate value for NaN cases
        else:
            start_time = datetime.combine(date.min, row['Start Time'])
            end_time = datetime.combine(date.min, row['End Time'])
            return (end_time - start_time).seconds / 3600

    availability_df['Working Hours'] = availability_df.apply(calculate_hours, axis=1)

    # Group by 'Practitioner ID', 'Staff Name', and 'User Role', then calculate the sum of 'Working Hours'
    total_hours_availability_df = availability_df.groupby(['Practitioner ID', 'Staff Name', 'User Role'])['Working Hours'].sum()

    # Reset the index to keep 'Practitioner ID', 'Staff Name', and 'User Role' as columns
    total_hours_availability_df = total_hours_availability_df.reset_index()

    return total_hours_availability_df










total_hours_availability_df = calculate_working_hours(availability_df)







#Create the pivot table to show exception block usage

exception_block_pivot_df = exception_appointment_df







# Create an availability_2_df to use with exception hours 
availability_2_df = availability_df







#Pull staff exceptions from NetSmart as schduled from exception definition

# Establish the connection
conn = pyodbc.connect(conn_stringPM)

# Define the SQL query for appt_staff_exceptions
sql_query_exceptions = """
SELECT *
FROM "SYSTEM".appt_staff_exceptions
WHERE exception_date >= ? AND exception_date <= ?
"""

# Read sql query into a DataFrame
appt_staff_exceptions_df = pd.read_sql(sql_query_exceptions, conn, params=[start_date, end_date])


# Define the SQL query for appt_staff_excep_definition
sql_query_definition = """
SELECT *
FROM "SYSTEM".appt_staff_excep_definition
WHERE exception_date <= ?
"""

# Read sql query into a DataFrame
appt_staff_exceptions_definition_df = pd.read_sql(sql_query_definition, conn, params=[end_date])


# Don't forget to close the connection
conn.close()






# Assuming appt_staff_exceptions_definition_df is your DataFrame
column_list = appt_staff_exceptions_definition_df.columns.tolist()




# Add the site_name value to the appt_staff_exceptions_df


# Create a function to extract the number from the 'HOL_uniqueid' column
def extract_number(id):
    return int(id.split('.')[1])

# Apply the function to the 'HOL_uniqueid' column
appt_staff_exceptions_df['HOL_uniqueid_number'] = appt_staff_exceptions_df['HOL_uniqueid'].apply(extract_number)

# Sort the dataframe by 'HOL_uniqueid_number'
sorted_df = appt_staff_exceptions_df.sort_values(by='HOL_uniqueid_number')

# Group by the necessary columns and get the first occurrence
grouped_df = sorted_df.groupby(['STAFFID', 'exception_description', 'entire_day_or_time_code', 'entire_day_or_time_value', 'exception_start_time', 'exception_end_time', 'data_entry_date', 'data_entry_by', 'data_entry_time', 'option_id']).first().reset_index()

# Merge the grouped dataframe with the definition dataframe to find the 'exception_site_name'
# Here we merge on multiple columns
merged_df = pd.merge(grouped_df, appt_staff_exceptions_definition_df, on=['STAFFID', 'data_entry_by', 'data_entry_date', 'data_entry_time', 'entire_day_or_time_value', 'exception_date', 'exception_start_time', 'exception_description'], how='left')

# Merge the original dataframe with the merged dataframe to append the 'exception_site_name'
appt_staff_exceptions_df = pd.merge(appt_staff_exceptions_df, merged_df[['STAFFID', 'exception_site_name']], on='STAFFID', how='left')





# Calculate correct exception time using the hoursw and exceptions form. Exceptions can be enters for ranges outside of availability so this 
# needs to be adjusted to only total exception time within the range of availability.

def convert_to_military_time(time_str):
    if isinstance(time_str, str):
        time_str = re.sub(r'(?<=\d)(?=[APap][Mm])', ' ', time_str)
        time_str = time_str.upper().strip()
        datetime_obj = dt.datetime.strptime(time_str, "%I:%M %p")
        return datetime_obj.strftime("%H:%M")
    return time_str

# Assuming the datetime columns might already be in correct format or need conversion from string
appt_staff_exceptions_df['exception_start_time'] = pd.to_datetime(
    appt_staff_exceptions_df['exception_start_time'].apply(convert_to_military_time), format='%H:%M', errors='coerce'
)
appt_staff_exceptions_df['exception_end_time'] = pd.to_datetime(
    appt_staff_exceptions_df['exception_end_time'].apply(convert_to_military_time), format='%H:%M', errors='coerce'
)

# Ensure 'Start Time' and 'End Time' are strings before concatenation
availability_2_df['Start Time'] = availability_2_df['Start Time'].astype(str)
availability_2_df['End Time'] = availability_2_df['End Time'].astype(str)

# Combine 'Date' with 'Start Time' and 'End Time' before converting to datetime
availability_2_df['Start Time'] = pd.to_datetime(
    availability_2_df['Date'].astype(str) + ' ' + availability_2_df['Start Time'], format='%Y-%m-%d %H:%M:%S', errors='coerce'
)
availability_2_df['End Time'] = pd.to_datetime(
    availability_2_df['Date'].astype(str) + ' ' + availability_2_df['End Time'], format='%Y-%m-%d %H:%M:%S', errors='coerce'
)

# Iterate over each row in appt_staff_exceptions_df
for index, row in appt_staff_exceptions_df.iterrows():
    staffid = row['STAFFID']
    exception_date = row['exception_date']
    start_time = row['exception_start_time']
    end_time = row['exception_end_time']
    entire_day_or_time_value = row['entire_day_or_time_value']
    exception_site_name = row.get('exception_site_name')  # Use .get() to avoid KeyError if column does not exist

    # Only consider records where exception_site_name is NULL
    if pd.isnull(exception_site_name):
        # Find matching rows in availability_2_df
        mask = (availability_2_df['Practitioner ID'] == staffid) & (availability_2_df['Date'] == exception_date)

        if entire_day_or_time_value == 'Selected Hours':
            # If the exception_start_time is before Start Time and the exception_end_time is before the End Time
            mask_update_start = mask & (availability_2_df['Start Time'] < start_time) & (availability_2_df['End Time'] > start_time)
            availability_2_df.loc[mask_update_start, 'Start Time'] = end_time

            # If the exception_start_time is after the Start Time and the exception_end_time is after the End Time
            mask_update_end = mask & (availability_2_df['Start Time'] < end_time) & (availability_2_df['End Time'] > end_time)
            availability_2_df.loc[mask_update_end, 'End Time'] = start_time

            # If the exception_start_time is before Start Time and the exception_end_time is after the End Time
            mask_remove = mask & (availability_2_df['Start Time'] >= start_time) & (availability_2_df['End Time'] <= end_time)
            availability_2_df = availability_2_df.loc[~mask_remove]

            # If the exception_start_time is after Start Time and the exception_end_time is before the End Time
            mask_split = mask & (availability_2_df['Start Time'] < start_time) & (availability_2_df['End Time'] > end_time)
            if mask_split.any():
                split_rows = availability_2_df.loc[mask_split].copy()
                split_rows['End Time'] = start_time
                availability_2_df.loc[mask_split, 'Start Time'] = end_time
                availability_2_df = pd.concat([availability_2_df, split_rows])

        elif entire_day_or_time_value == 'Entire Day':
            # Remove rows where entire_day_or_time_value = ‘Entire Day’
            availability_2_df = availability_2_df.loc[~mask]

# Remove rows where 'Start Time' is not before 'End Time'
availability_2_df = availability_2_df[availability_2_df['Start Time'] < availability_2_df['End Time']]







# adjust availability if exception hours has SITE data

# Explode the appt_staff_exceptions_df on the exception_site_name column
appt_staff_exceptions_df = appt_staff_exceptions_df.assign(exception_site_name=appt_staff_exceptions_df['exception_site_name'].str.split('&')).explode('exception_site_name')

# # Filter out exceptions where exception_site_name is NOT in available_time_sites
# appt_staff_exceptions_df = appt_staff_exceptions_df[
#     appt_staff_exceptions_df['exception_site_name'].isin(available_time_sites)
# ]


# Iterate over each row in appt_staff_exceptions_df
for index, row in appt_staff_exceptions_df.iterrows():
    staffid = row['STAFFID']
    exception_date = row['exception_date']
    start_time = row['exception_start_time']
    end_time = row['exception_end_time']
    entire_day_or_time_value = row['entire_day_or_time_value']
    exception_site_name = row['exception_site_name']  # assuming this column exists in your dataframe

    # Only consider records where exception_site_name is NOT NULL
    if pd.notnull(exception_site_name):
        # Find matching rows in availability_2_df
        mask = (availability_2_df['Practitioner ID'] == staffid) & (availability_2_df['Date'] == exception_date)

        # If the exception_start_time is before Start Time and the exception_end_time is before the End Time
        mask_update_start = mask & (availability_2_df['Start Time'] < start_time) & (availability_2_df['End Time'] > start_time)
        availability_2_df.loc[mask_update_start, 'Start Time'] = end_time

        # If the exception_start_time is after the Start Time and the exception_end_time is after the End Time
        mask_update_end = mask & (availability_2_df['Start Time'] < end_time) & (availability_2_df['End Time'] > end_time)
        availability_2_df.loc[mask_update_end, 'End Time'] = start_time

        # If the exception_start_time is before Start Time and the exception_end_time is after the End Time
        mask_remove = mask & (availability_2_df['Start Time'] >= start_time) & (availability_2_df['End Time'] <= end_time)
        availability_2_df = availability_2_df.loc[~mask_remove]

        # If the exception_start_time is after Start Time and the exception_end_time is before the End Time
        mask_split = mask & (availability_2_df['Start Time'] < start_time) & (availability_2_df['End Time'] > end_time)
        if mask_split.any():
            split_rows = availability_2_df.loc[mask_split].copy()
            split_rows['End Time'] = start_time
            availability_2_df.loc[mask_split, 'Start Time'] = end_time
            availability_2_df = pd.concat([availability_2_df, split_rows])

        # For records where the entire_day_or_time_value = ‘Entire Day’, remove the record from the availability_2_df
        if entire_day_or_time_value == 'Entire Day':
            availability_2_df = availability_2_df.loc[~mask]

# Remove rows where 'Start Time' is not before 'End Time'
availability_2_df = availability_2_df[availability_2_df['Start Time'] < availability_2_df['End Time']]







#ensure not too much system exception time is taken. cannot be more than the working day available time.

# Convert date columns in DataFrames to datetime.date
availability_df['Date'] = pd.to_datetime(availability_df['Date'], errors='coerce').dt.date
availability_2_df['Date'] = pd.to_datetime(availability_2_df['Date'], errors='coerce').dt.date
exception_appointment_df['appointment_date'] = pd.to_datetime(exception_appointment_df['appointment_date'], errors='coerce').dt.date


# Ensure start_date and end_date are of type datetime.date
if isinstance(start_date, str) and start_date:
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
if isinstance(end_date, str) and end_date:
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

# Filter availability_df
mask_availability = (availability_df['Date'] >= start_date) & (availability_df['Date'] <= end_date)
availability_df = availability_df.loc[mask_availability]

# Filter availability_2_df
mask_availability_2 = (availability_2_df['Date'] >= start_date) & (availability_2_df['Date'] <= end_date)
availability_2_df = availability_2_df.loc[mask_availability_2]

# Filter exception_appointment_df
mask_exception_time = (exception_appointment_df['appointment_date'] >= start_date) & (exception_appointment_df['appointment_date'] <= end_date)
exception_appointment_df = exception_appointment_df.loc[mask_exception_time]









# Check if 'duration_minutes' exists in the DataFrame
if 'duration_minutes' in exception_appointment_df.columns:
    # Convert 'duration_minutes' to numeric type before grouping
    exception_appointment_df['duration_minutes'] = pd.to_numeric(exception_appointment_df['duration_minutes'], errors='coerce')

    # Group by STAFFID and calculate the total duration_minutes
    total_duration_per_staff = exception_appointment_df.groupby('STAFFID')['duration_minutes'].sum()

    # Convert total duration from minutes to hours and round to two decimal places
    total_duration_per_staff = total_duration_per_staff.apply(lambda x: round(x / 60, 2))

    # Reset the index and rename the columns
    exception_time_df = total_duration_per_staff.reset_index().rename(columns={'duration_minutes': 'Total_Appt_Exception_Hours'})

else:
    print("Column 'duration_minutes' does not exist in exception_appointment_df")








# Calculate total ADJUSTED available working hours for each Practitioner ID in availability_2_df
availability_2_df['Working Hours'] = (availability_2_df['End Time'] - availability_2_df['Start Time']).dt.total_seconds() / 3600








# total the working hours AFTER system exception time has been removed

# Define function to calculate total working hours from the given dataframe
def calculate_total_working_hours(availability_2_df):
    # Group by 'Practitioner ID', 'Staff Name', and 'User Role', then calculate the sum of 'Working Hours'
    total_hours_availability_df = availability_2_df.groupby(['Practitioner ID', 'Staff Name', 'User Role'])['Working Hours'].sum()

    # Reset the index to keep 'Practitioner ID', 'Staff Name', and 'User Role' as columns
    total_hours_availability_df = total_hours_availability_df.reset_index()

    return total_hours_availability_df

# Calculate total working hours
availability_2_df = calculate_total_working_hours(availability_2_df)










# # Calculating the difference in Working Hours
# difference_df = pd.DataFrame({
#     "Practitioner ID": total_hours_availability_df["Practitioner ID"],
#     "Difference in Working Hours": total_hours_availability_df["Working Hours"] - availability_2_df["Working Hours"]
# })


before = total_hours_availability_df.set_index('Practitioner ID')['Working Hours']
after  = availability_2_df.set_index('Practitioner ID')['Working Hours']

diff_series = (before - after).fillna(0).rename('Difference in Working Hours')
difference_df = diff_series.reset_index()






# combine the system exceptions and the appt exception blocks into one totalin difference_df

difference_df['Practitioner ID'] = difference_df['Practitioner ID'].astype(str)
exception_time_df['STAFFID'] = exception_time_df['STAFFID'].astype(str)

# Remove leading zeros from STAFFID
exception_time_df['STAFFID'] = exception_time_df['STAFFID'].str.lstrip('0')

# Also remove leading zeros from Practitioner ID for consistency
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].str.lstrip('0')

# Merge the two dataframes on Practitioner ID and STAFFID
merged_df = pd.merge(difference_df, exception_time_df, left_on='Practitioner ID', right_on='STAFFID', how='inner')

# Add the two columns together
merged_df['Total Hours'] = merged_df['Difference in Working Hours'] + merged_df['Total_Appt_Exception_Hours']

# Drop unnecessary columns if needed
difference_df = merged_df.drop(columns=['STAFFID'])









adjusted_total_hours_availability_df = availability_2_df.groupby('Practitioner ID')['Working Hours'].sum()

# Store the total adjusted hours in a variable for use in a report later
adjusted_total_hours_availability_df = adjusted_total_hours_availability_df.to_dict()






exception_time_df = exception_time_df.rename(columns={'STAFFID': 'Practitioner ID'})







#Adding the appointment exception and the system exception time together

# Convert Practitioner ID to string and pad with leading zeros
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].astype(str).str.zfill(6)
exception_time_df['Practitioner ID'] = exception_time_df['Practitioner ID'].astype(str).str.zfill(6)

# Merging dataframes on 'Practitioner ID'
merged_df = exception_time_df.merge(difference_df[['Practitioner ID', 'Difference in Working Hours']], on='Practitioner ID', how='left')

# Replacing NaN values in 'Difference in Working Hours' with 0
merged_df['Difference in Working Hours'] = merged_df['Difference in Working Hours'].fillna(0)

# Adding 'Difference in Working Hours' to 'Total_Appt_Exception_Hours'
merged_df['Total_Appt_Exception_Hours'] += merged_df['Difference in Working Hours']

# Dropping the 'Difference in Working Hours' column if not needed anymore
difference_df = merged_df.drop(columns=['Difference in Working Hours'])








#Total working hours for each practitioner
# Group by 'Practitioner ID' and 'Site', then sum the 'Working Hours' for each group
availability_df = availability_df.groupby(['Practitioner ID', 'Site'], as_index=False)['Working Hours'].sum()








# Append desired user_roles to the STAFFID as specified in the config.ini

user_roles = user_roles.split(',')

# Create a new connection
conn = pyodbc.connect(conn_stringPM)

# Create a SQL query to select all data from the SYSTEM.RADplus_users table
sql_query = "SELECT staff_member_id AS STAFFID, USERROLE FROM SYSTEM.RADplus_users"

# Execute the SQL query and store the result in a DataFrame
all_data_df = pd.read_sql(sql_query, conn)

conn.close()







# Filter rows where USERROLE contains 'PRESCRIBER'
df_filtered = all_data_df[all_data_df['USERROLE'].str.contains('PRESCRIBER', na=False)]







# Step 1 and 2: Iterate through the DataFrame
filtered_data = []
for index, row in all_data_df.iterrows():
    userrole = row['USERROLE']
    if userrole is not None:
        # Remove leading and trailing '&'
        userrole = userrole.strip('&')
        # Split the USERROLE by '&'
        roles = userrole.split('&')
        # Check if the record has more than one user role from the user_roles parameter
        matched_roles = [role for role in roles if role in user_roles]
        if matched_roles:
            # If the record has one or more user roles from the user_roles parameter, keep it
            row['USERROLE'] = ','.join(matched_roles)  # Replace '&' with ','
            filtered_data.append(row)

# Step 3: Create a new DataFrame with the filtered and formatted records
all_data_df = pd.DataFrame(filtered_data)





# Reorder the columns
total_hours_availability_df = total_hours_availability_df.reindex(columns=['Staff Name', 'Practitioner ID', 'Working Hours'])







# Ensure 'Practitioner ID' in both DataFrames are treated as strings
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].astype(str)
total_hours_availability_df['Practitioner ID'] = total_hours_availability_df['Practitioner ID'].astype(str)

# Remove leading zeros from 'Practitioner ID' in difference_df
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].str.lstrip('0')


# Merge the DataFrames on 'Practitioner ID'
total_hours_availability_df = total_hours_availability_df.merge(difference_df[['Practitioner ID', 'Total_Appt_Exception_Hours']], on='Practitioner ID', how='left')

# Rename 'Total_Appt_Exception_Hours' to 'TOTAL EXCEPTION TIME' in the resulting DataFrame
total_hours_availability_df = total_hours_availability_df.rename(columns={'Total_Appt_Exception_Hours': 'Total Exception Hours'})







# Ensure 'Practitioner ID' in both DataFrames are treated as strings
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].astype(str)
total_hours_availability_df['Practitioner ID'] = total_hours_availability_df['Practitioner ID'].astype(str)

# Remove leading zeros from 'Practitioner ID' in difference_df
difference_df['Practitioner ID'] = difference_df['Practitioner ID'].str.lstrip('0')

# Merge the DataFrames on 'Practitioner ID'
merged_df = total_hours_availability_df.merge(difference_df[['Practitioner ID', 'Total_Appt_Exception_Hours']], on='Practitioner ID', how='left')

# Create the 'Adjusted Available Time' column
merged_df['Adjusted Available Hours'] = merged_df['Working Hours'] - merged_df['Total_Appt_Exception_Hours']

# Add the new column back to the original DataFrame
total_hours_availability_df['Adjusted Available Hours'] = merged_df['Adjusted Available Hours']






# Check if 'PRESCRIBER' exists in the 'USERROLE' column
all_data_df['contains_PRESCRIBER'] = all_data_df['USERROLE'].apply(lambda x: 'PRESCRIBER' in x if x is not None else False)

# Filter the DataFrame to only include rows where 'USERROLE' contains 'PRESCRIBER'
prescriber_df = all_data_df[all_data_df['contains_PRESCRIBER'] == True]









if use_note_table == "1" and use_note_billing_charge_table == "0":
    # Establish a connection to the database
    df_csv = pd.read_csv(productivity_service_code_list_location)
    conn = pyodbc.connect(conn_stringPM)
    
    # Define the first SQL query
    sql_query1 = """
        SELECT DISTINCT SERVICE_CODE, cpt_code, charge, duration_range 
        FROM SYSTEM.billing_tx_master_fee_table 
    """
    df1 = pd.read_sql(sql_query1, conn)
    
    df1[['duration_range_start', 'duration_range_end']] = df1['duration_range'].str.split('-', expand=True)
    df1[['duration_range_start', 'duration_range_end']] = df1[['duration_range_start', 'duration_range_end']].apply(pd.to_numeric)
    
    # Ensure that draft_final_code and document_routing_status are formatted as tuples of strings
    draft_final_code_tuple = tuple(draft_final_code.split(","))
    document_routing_status_tuple = tuple(document_routing_status.split(","))
    
    draft_final_code_str = ', '.join(f"'{code.strip()}'" for code in draft_final_code_tuple)
    document_routing_status_str = ', '.join(f"'{status.strip()}'" for status in document_routing_status_tuple)
    
    # Define a function to validate the date parameters
    def validate_date_parameters(use_date_of_service, use_date_of_note):
        if use_date_of_service == "1" and use_date_of_note == "1":
            print("Please check config file [\"productive_time\"][\"use_date_of_service\"] AND [\"productive_time\"][\"use_date_of_note\"]")
            return False
        elif use_date_of_service != "1" and use_date_of_note != "1":
            print("Neither use_date_of_service nor use_date_of_note is set to 1.")
            return False
        return True
        
    # Validate the date parameters
    if not validate_date_parameters(use_date_of_service, use_date_of_note):
        raise ValueError("Invalid configuration: both use_date_of_service and use_date_of_note cannot be set to 1 simultaneously.")
    
    # Determine the date column to use based on the parameters
    date_column = 'date_of_service' if use_date_of_service == "1" else 'date_of_note' if use_date_of_note == "1" else None
    
    # Raise an error if no valid date column is specified
    if date_column is None:
        raise ValueError("Invalid configuration: neither use_date_of_service nor use_date_of_note is set to 1.")
    
    # Calculate the Monday after the end_date at 8:00 AM
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    if end_date.weekday() == 0:  # If end_date is a Monday
        next_monday = end_date + timedelta(days=7)
    else:
        next_monday = end_date + timedelta(days=(7 - end_date.weekday()))
    next_monday_8am = datetime.combine(next_monday, datetime.min.time()).replace(hour=8, minute=0, second=0)

    # Convert next_monday_8am to string in the required format
    next_monday_date_str = next_monday.strftime('%Y-%m-%d')
    next_monday_8am_str = next_monday_8am.strftime('%I:%M %p')

    # Define the second SQL query with the additional condition
    sql_query2 = f"""
        SELECT 
            cpn.service_duration,
            cpn.date_of_note, 
            cpn.draft_final_code, 
            cpn.practitioner_id, 
            cpn.service_charge_code,
            cpn.data_entry_date,
            cpn.data_entry_time
        FROM 
            CWSSYSTEM.cw_patient_notes AS cpn
        WHERE 
            cpn.{date_column} >= '{start_date}' AND cpn.{date_column} <= '{end_date}'
            AND cpn.draft_final_code IN ({draft_final_code_str})
            AND cpn.document_routing_status IN ({document_routing_status_str})
            AND (cpn.data_entry_date < '{next_monday_date_str}' OR (cpn.data_entry_date = '{next_monday_date_str}' AND cpn.data_entry_time <= '{next_monday_8am_str}'))
    """

    
    # Execute the query and fetch the results
    df2 = pd.read_sql(sql_query2, conn)
    
    # Close the connection
    conn.close()
    
    # Fill NaN values in df2 and df1 with 0 for service duration
    df2['service_duration'] = df2['service_duration'].fillna(0).astype(int)
    df1['duration_range_start'] = df1['duration_range_start'].replace([np.inf, -np.inf], np.nan).fillna(0).astype(int)
    df1['duration_range_end'] = df1['duration_range_end'].replace([np.inf, -np.inf], np.nan).fillna(0).astype(int)
    
    # Define the apply_logic function
    def apply_logic(row):
        # Get the matching rows from df1
        matching_rows = df1[df1['SERVICE_CODE'] == row['service_charge_code']]
        
        # If SERVICE_CODE == cpt_code, return cpt_code
        if not matching_rows.empty and matching_rows['SERVICE_CODE'].equals(matching_rows['cpt_code']):
            return matching_rows['cpt_code'].values[0]
        
        # If service_charge_code ends with "F" or "TH", remove "F" or "TH" and return
        elif row['service_charge_code'] is not None and row['service_charge_code'].endswith(('F', 'TH')):
            return row['service_charge_code'].rstrip('F').rstrip('TH')
        
        # If SERVICE_CODE != cpt_code, find the cpt_code where service_duration is within the range
        elif not matching_rows.empty:
            duration_match = matching_rows[(matching_rows['duration_range_start'] <= row['service_duration']) & (matching_rows['duration_range_end'] >= row['service_duration'])]
            if not duration_match.empty:
                return duration_match['cpt_code'].values[0]
            elif row['service_charge_code'] == '9083A':
                return '90833'
        
        # Special case for service_charge_code '9083X'
        if row['service_charge_code'] == '9083X':
            return '90832'

        # Return None if no match found
        return None
    
    # Apply the function to df2
    df2['cpt_code'] = df2.apply(apply_logic, axis=1)
    
    # Check for service_charge_code longer than 5 characters and cpt_code is None
    mask = (df2['service_charge_code'].str.len() > 5) & (df2['cpt_code'].isnull())
    
    # Process the service_charge_code to strip after 5 characters and assign to cpt_code
    df2.loc[mask, 'cpt_code'] = df2.loc[mask, 'service_charge_code'].str[:5]
    
    # Additional logic to handle cases where service_charge_code is 5 characters or less and cpt_code is still None
    mask_short = (df2['service_charge_code'].str.len() <= 5) & (df2['cpt_code'].isnull())
    
    # Assign service_charge_code to cpt_code
    df2.loc[mask_short, 'cpt_code'] = df2.loc[mask_short, 'service_charge_code']
    
    # Check for any remaining null or zero values in cpt_code (if needed)
    null_values = df2[df2['cpt_code'].isnull()]
    zero_values = df2[df2['cpt_code'] == 0]
    
    # Rename columns
    df2.rename(columns={'practitioner_id': 'STAFFID', 'cpt_code': 'SERVICE_CODE'}, inplace=True)
    
    # Drop all other columns
    columns_to_keep = ['STAFFID', 'SERVICE_CODE']
    df2 = df2[columns_to_keep]

    # Get the unique cpt_codes from df_csv
    cpt_codes = df_csv['cpt_code'].unique()
    
    # Keep only the records in df2 where SERVICE_CODE is in cpt_codes
    schedule_df = df2[df2['SERVICE_CODE'].isin(cpt_codes)]

elif use_note_billing_charge_table == "1" and use_note_table == "0":
    # Pull in the schedule_df2 from the staff_tx_history table for billed services.

    # Create a new connection
    conn = pyodbc.connect(conn_stringPM)
    
    # Create a new cursor
    cursor = conn.cursor()
    
    # SQL query to pull records from the staff_tx_history table
    sql_query1 = f"""
    SELECT v_PROVIDER_ID AS STAFFID, v_SERVICE_CODE AS service_code, guarantor_liability AS cost_of_service
    FROM SYSTEM.billing_tx_charge_detail
    WHERE data_entry_date >= '{schedule_start_date}' AND data_entry_date <= '{schedule_end_date}'
    """
    
    df1 = pd.read_sql(sql_query1, conn)
    
    # Close the connection
    conn.close()

    df_csv = pd.read_csv(productivity_service_code_list_location)
    # Perform the merge operation
    schedule_df = pd.merge(df1, df_csv[['service_code', 'cost_of_service', 'cpt_code']], how='left', on=['service_code', 'cost_of_service'])

    # Drop the 'SERVICE_CODE' column
    schedule_df = schedule_df.drop(['service_code','cost_of_service'], axis=1)
    
    # Rename the 'cpr_code' column to 'SERVICE_CODE'
    schedule_df = schedule_df.rename(columns={'cpt_code': 'SERVICE_CODE'})

elif use_note_table == use_note_billing_charge_table:
    print("Invalid table to query Productive Time Services. Please check [\"productive_time\"][\"use_note_table\"] and [\"productive_time\"][\"use_note_billing_charge_table\"]")
else:
    print("No valid condition met.")






# Group by STAFFID and SERVICE_CODE, then count occurrences
service_code_counts = schedule_df.groupby(['STAFFID', 'SERVICE_CODE']).size().reset_index(name='COUNT')






# Drop records from service_code_counts that are not eligible service codes

productivity_service_code_list = pd.read_csv(productivity_service_code_list_location)






# Drop records from service_code_counts that are not eligible service codes

# Select only the relevant columns from productivity_service_code_list
productivity_service_code_list = productivity_service_code_list[['cpt_code', 'value', 'role']]

# Merge the dataframes on the SERVICE_CODE column from service_code_counts and the cpt_code column from productivity_service_code_list
merged_df = pd.merge(service_code_counts, productivity_service_code_list, left_on='SERVICE_CODE', right_on='cpt_code', how='left')

# Create the new column by multiplying 'COUNT' by 'value'
merged_df['Total Productivity Hours'] = merged_df['COUNT'] * merged_df['value']

# Keep only the rows where 'value' is not null
service_code_counts = merged_df[merged_df['value'].notna()]







# Ensure 'Practitioner ID' is of type object in both dataframes
total_hours_availability_df['Practitioner ID'] = total_hours_availability_df['Practitioner ID'].astype(str)
availability_2_df['Practitioner ID'] = availability_2_df['Practitioner ID'].astype(str)

# Merging the dataframes on 'Practitioner ID'
merged_df = total_hours_availability_df.merge(availability_2_df[['Practitioner ID', 'User Role']], on='Practitioner ID', how='left')

# Renaming the 'User Role' column to 'USERROLE'
total_hours_availability_df = merged_df.rename(columns={'User Role': 'USERROLE'})






total_hours_availability_df_2 = total_hours_availability_df.copy()







# Remove leading zeros from STAFFID
service_code_counts['STAFFID'] = service_code_counts['STAFFID'].astype(str).str.lstrip('0')

# Merge the dataframes
total_hours_availability_df = total_hours_availability_df.merge(service_code_counts, left_on='Practitioner ID', right_on='STAFFID', how='left')







# Filter rows where the 'role' column matches the 'USERROLE' column
total_hours_availability_df = total_hours_availability_df[total_hours_availability_df['role'] == total_hours_availability_df['USERROLE']]






# Pivot the SERVICE_CODE column
transposed_df = total_hours_availability_df.pivot_table(index='Practitioner ID', columns='SERVICE_CODE', values='COUNT').fillna(0)

# Reset index to merge easily
transposed_df = transposed_df.reset_index()

# Merge the original dataframe with the transposed dataframe
total_hours_availability_df = pd.merge(total_hours_availability_df.drop(columns=['SERVICE_CODE', 'COUNT']), transposed_df, on='Practitioner ID', how='left').drop_duplicates()




# 1. SUM all values for each 'Practitioner ID' in the 'Total Productivity Hours' column and put the value in a new column called 'Productivity Hours'
total_hours_availability_df['Productivity Hours'] = total_hours_availability_df.groupby('Practitioner ID')['Total Productivity Hours'].transform('sum')

# 2. Delete the following columns: STAFFID, cpt_code, value, role
total_hours_availability_df = total_hours_availability_df.drop(columns=['STAFFID', 'cpt_code', 'value', 'role', 'Total Productivity Hours'])

# 3. Drop duplicate records
total_hours_availability_df = total_hours_availability_df.drop_duplicates()





# Add a new column 'Productivity Percentage' that is the result of the division of 'Productivity Hours' by 'Adjusted Available Time'
total_hours_availability_df['Productivity Percentage'] = total_hours_availability_df['Productivity Hours'] / total_hours_availability_df['Adjusted Available Hours']

total_hours_availability_df['Productivity Percentage'] = (total_hours_availability_df['Productivity Percentage'] * 100).round(2)






# Create a new column 'Actual Productivity Percentage' with rounded values
total_hours_availability_df['Actual Productivity Percentage'] = total_hours_availability_df['Productivity Percentage'].round()

# Rename the original 'Productivity Percentage' column to preserve it
total_hours_availability_df.rename(columns={'Productivity Percentage': 'Original Productivity Percentage'}, inplace=True)

# Create a new 'Productivity Percentage' with the rounded values
total_hours_availability_df['Productivity Percentage'] = total_hours_availability_df['Actual Productivity Percentage']







# Drop the 'Actual Productivity Percentage' column
total_hours_availability_df.drop('Actual Productivity Percentage', axis=1, inplace=True)

# Define the order for the primary columns as specified
primary_columns = ['Staff Name', 'Practitioner ID', 'Working Hours', 'Total Exception Hours', 
                   'Adjusted Available Hours', 'Productivity Hours', 'Original Productivity Percentage', 'Productivity Percentage']

# Include any additional columns that were not listed in primary_columns
additional_columns = [col for col in total_hours_availability_df.columns if col not in primary_columns]

# Combine the primary columns with the additional columns for the new order
new_column_order = primary_columns + additional_columns

# Reassign the DataFrame with the new column order
total_hours_availability_df = total_hours_availability_df[new_column_order]






import numpy as np

# Read the CSV file into a DataFrame
payout_df = pd.read_csv(productivity_payout_percentage_list_location)

# Initialize a new column 'Productivity Payout' with NaN values
total_hours_availability_df['Productivity Payout'] = np.nan

# Iterate over the rows of 'total_hours_availability_df'
for i in total_hours_availability_df.index:
    # Get the 'Productivity Percentage' and 'USERROLE' for the current row
    prod_percent = total_hours_availability_df.loc[i, 'Productivity Percentage']
    user_role = total_hours_availability_df.loc[i, 'USERROLE']
    
    # Filter payout_df to include only rows with the same role
    role_specific_payout_df = payout_df[payout_df['Role'] == user_role]
    
    # Check if 'Productivity Percentage' is above 100%
    if prod_percent > 100:
        # Find the row in role_specific_payout_df corresponding to 100% productivity
        max_payout_rows = role_specific_payout_df[role_specific_payout_df['Upper Percentile'] == 100]
        if not max_payout_rows.empty:
            max_payout_row = max_payout_rows.iloc[0]
            total_hours_availability_df.loc[i, 'Productivity Payout'] = max_payout_row['Productivity Dollar Payout']
    else:
        # Iterate over the rows of 'role_specific_payout_df'
        for j in role_specific_payout_df.index:
            # Check if 'Productivity Percentage' is within the current percentile range
            if prod_percent >= role_specific_payout_df.loc[j, 'Lower Percentile'] and prod_percent <= role_specific_payout_df.loc[j, 'Upper Percentile']:
                # If it is, set the 'Productivity Dollar Payout' for the current row in 'total_hours_availability_df'
                total_hours_availability_df.loc[i, 'Productivity Payout'] = role_specific_payout_df.loc[j, 'Productivity Dollar Payout']
                break







# Specify the columns you want at the beginning
cols_to_order = ['Staff Name', 'Practitioner ID', 'Working Hours', 'Total Exception Hours', 'Adjusted Available Hours', 'USERROLE', 'Productivity Hours', 'Original Productivity Percentage', 'Productivity Percentage', 'Productivity Payout']
# Create a list of the other columns (those not specified above) in their current order
other_cols = [col for col in total_hours_availability_df.columns if col not in cols_to_order]

# Concatenate the two lists to get a new column order
new_order = cols_to_order + other_cols

# Rearrange the columns
total_hours_availability_df = total_hours_availability_df[new_order]

# Rename 'Working Hours' to 'Available Time'
total_hours_availability_df.rename(columns={'Working Hours': 'Available Time'}, inplace=True)

# Convert 'Productivity Percentage' and 'Productivity Payout' to string and add '%' and '$'
total_hours_availability_df['Productivity Percentage'] = total_hours_availability_df['Productivity Percentage'].astype(str) + '%'
total_hours_availability_df['Productivity Payout'] = '$' + total_hours_availability_df['Productivity Payout'].astype(str)

total_hours_availability_df['Original Productivity Percentage'] = total_hours_availability_df['Original Productivity Percentage'].astype(str) + '%'
total_hours_availability_df['Productivity Payout'] = '$' + total_hours_availability_df['Productivity Payout'].astype(str)








draft_df = total_hours_availability_df.copy()





#Formatting the dataframe for EXCEL



# 2) Set values to 2 decimal places, replace NaN values with 0
cols_to_format = ['Available Time', 'Total Exception Hours', 'Adjusted Available Hours', 'Productivity Hours']
for col in cols_to_format:
    draft_df[col] = draft_df[col].replace(np.nan, 0).round(2)

# 3) Replace 'nan%' with '0' in 'Productivity Percentage' column
draft_df['Original Productivity Percentage'] = draft_df['Original Productivity Percentage'].replace('nan%', '0')
draft_df['Productivity Percentage'] = draft_df['Productivity Percentage'].replace('nan%', '0')

# 4) Process the 'Productivity Payout' column
# Define a function to clean and convert the values
def clean_payout_value(value):
    if pd.isna(value):
        return 0
    elif isinstance(value, str):
        # Remove any dollar signs or commas and convert to float
        value = value.replace('$', '').replace(',', '')
        try:
            return float(value)
        except ValueError:
            return 0
    elif isinstance(value, (int, float)):
        return value
    else:
        return 0

# Apply the function to the 'Productivity Payout' column
draft_df['Productivity Payout'] = draft_df['Productivity Payout'].apply(clean_payout_value)

# Replace NaN values with 0 (this is a safeguard step; ideally should be unnecessary)
draft_df['Productivity Payout'] = draft_df['Productivity Payout'].fillna(0)

# Add the dollar sign back to the numbers
draft_df['Productivity Payout'] = draft_df['Productivity Payout'].apply(
    lambda x: f"${x:.2f}" if isinstance(x, (int, float)) and not pd.isna(x) else x
)

# Convert the 'Practitioner ID', 'Productivity Percentage', and 'Productivity Payout' columns to string type
draft_df['Practitioner ID'] = draft_df['Practitioner ID'].astype(str)
draft_df['Productivity Percentage'] = draft_df['Productivity Percentage'].astype(str)
draft_df['Productivity Payout'] = draft_df['Productivity Payout'].astype(str)






# Sort the dataframe by 'Staff Name' in alphabetical order
draft_df = draft_df.sort_values('Staff Name')






# Create selected_draft_df by extracting only the 'Staff Name' and 'Practitioner ID' columns from draft_df
selected_draft_df = draft_df[['Staff Name', 'Practitioner ID']].copy()

# Pad the 'Practitioner ID' column with leading zeros to match the length of 'STAFFID'
selected_draft_df['Practitioner ID'] = selected_draft_df['Practitioner ID'].astype(str).str.zfill(6)

# Create the pivot table
count_df = exception_block_pivot_df.groupby(['STAFFID', 'SERVICE_CODE']).size().reset_index(name='COUNT')
pivot_df = count_df.pivot(index='STAFFID', columns='SERVICE_CODE', values='COUNT')

# Replace NaN values with 0 and convert to integer
pivot_df = pivot_df.fillna(0).astype(int).reset_index()

# Pad the 'Practitioner ID' column with leading zeros to match the length of 'STAFFID'
selected_draft_df['Practitioner ID'] = selected_draft_df['Practitioner ID'].astype(str).str.zfill(6)

# Perform the merge
merged_df = pivot_df.merge(selected_draft_df, left_on='STAFFID', right_on='Practitioner ID', how='inner')

# Drop the 'STAFFID' column if it exists
if 'STAFFID' in merged_df.columns:
    merged_df = merged_df.drop(columns=['STAFFID'])

# Reorder the columns to have 'Practitioner ID' first
cols = ['Practitioner ID'] + [col for col in merged_df if col != 'Practitioner ID']
merged_df = merged_df[cols]

# Sort the dataframe by 'Staff Name' in alphabetical order
merged_df = merged_df.sort_values('Staff Name')

# Reorder the columns to have 'Staff Name' as the first column
cols = ['Staff Name'] + [col for col in merged_df if col != 'Staff Name']
merged_df = merged_df[cols]

# Replace all cells that have a 0 in them with an empty string
exception_block_df = merged_df.replace(0, '')




# Extract the necessary columns
draft_df_extracted = draft_df[['Practitioner ID', 'Productivity Payout']]
staff_working_days_hours_extracted = staff_working_days_hours[['STAFFID', 'FTE']]

# Ensure the IDs are strings for proper matching
draft_df_extracted['Practitioner ID'] = draft_df_extracted['Practitioner ID'].astype(str)
staff_working_days_hours_extracted['STAFFID'] = staff_working_days_hours_extracted['STAFFID'].astype(str)

# Remove leading zeros from STAFFID for matching purposes
staff_working_days_hours_extracted['STAFFID'] = staff_working_days_hours_extracted['STAFFID'].str.lstrip('0')

# Merge the DataFrames based on Practitioner ID and STAFFID
merged_df = draft_df_extracted.merge(staff_working_days_hours_extracted, left_on='Practitioner ID', right_on='STAFFID')

# Drop the STAFFID column if it's no longer needed
merged_df.drop('STAFFID', axis=1, inplace=True)

# Remove the dollar sign and convert Productivity Payout to float
merged_df['Productivity Payout'] = merged_df['Productivity Payout'].replace('[\$,]', '', regex=True).astype(float)

# Ensure FTE is of type float
merged_df['FTE'] = merged_df['FTE'].astype(float)

# Drop duplicate records
merged_df.drop_duplicates(inplace=True)

# Multiply Productivity Payout by FTE
merged_df['Adjusted Productivity Payout'] = merged_df['Productivity Payout'] * merged_df['FTE']

# Merge the adjusted productivity payout back into the original draft_df
draft_df = draft_df.merge(merged_df[['Practitioner ID', 'Adjusted Productivity Payout']], on='Practitioner ID', how='left')

# Update the Productivity Payout in draft_df
draft_df['Productivity Payout'] = draft_df['Adjusted Productivity Payout']

# Drop the Adjusted Productivity Payout column
draft_df.drop('Adjusted Productivity Payout', axis=1, inplace=True)

# Add a '$' in front of each value in the Productivity Payout column
draft_df['Productivity Payout'] = draft_df['Productivity Payout'].apply(lambda x: f"${x:.2f}")




# Create the legend for the STAFF CODES

# Establish the connection
conn = pyodbc.connect(conn_stringPM)

# Define the SQL query
sql_query = """
SELECT DISTINCT(SERVICE_CODE), service_description
FROM SYSTEM.AppointmentData 
WHERE SERVICE_CODE LIKE '%STAFF%'
"""

# Execute the query and assign the result to a pandas DataFrame
scheduler_exception_legend_df = pd.read_sql(sql_query, conn)

# Close the connection
conn.close()








# Create the Excel File

import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email import encoders
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
import openpyxl


# Write the DataFrame to an Excel file with auto-adjusted column width
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    for i, row in draft_df.iterrows():
        if i == 0:
            draft_df.to_excel(writer, sheet_name='Productivity', startrow=4, startcol=1, header=True, index=False)
        row_df = row.to_frame().transpose()
        row_df.to_excel(writer, sheet_name=row['Staff Name'], startrow=4, startcol=1, header=True, index=False)
    
    # Write the exception_block_df DataFrame to the 'Scheduler_Exceptions' sheet
    exception_block_df.to_excel(writer, sheet_name='Scheduler_Exceptions', startrow=4, startcol=1, header=True, index=False)

    # Get the number of columns in exception_block_df
    num_cols_exception_block = exception_block_df.shape[1]

    # Write the scheduler_exception_legend_df DataFrame to the 'Scheduler_Exceptions' sheet, a couple of columns to the right of exception_block_df
    scheduler_exception_legend_df.to_excel(writer, sheet_name='Scheduler_Exceptions', startrow=4, startcol=num_cols_exception_block + 3, header=True, index=False)

# Load the workbook and process each sheet
book = openpyxl.load_workbook('output.xlsx')
for sheet_name in ['Productivity'] + draft_df['Staff Name'].tolist():
    sheet = book[sheet_name]
    
    # Adjust column widths for columns B - K
    for col in range(2, 12):  # Columns B - K
        max_length = 0
        column = chr(64 + col)  # Convert to column letter
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Right align all columns except for column B
    for col in range(3, sheet.max_column + 1):  # Start from column C
        for cell in sheet[chr(64 + col)]:
            cell.alignment = Alignment(horizontal='right')

    # Left align column B
    for cell in sheet['B']:
        cell.alignment = Alignment(horizontal='left')

    # Make headers bold and center aligned, and remove borders
    for cell in sheet[5]:  # The headers are in row 5
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(border_style=None),
                             right=Side(border_style=None),
                             top=Side(border_style=None),
                             bottom=Side(border_style=None))

    # Remove grid lines
    sheet.sheet_view.showGridLines = False

    # Define border styles
    bottom_border = Border(bottom=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))
    left_border = Border(left=Side(style='thin'))

    # Apply common formatting and logo for all sheets
    title = f'Productivity Report {start_date} to {end_date}'
    sheet.insert_rows(2)
    sheet['B2'] = title

    # Load the image
    img = Image('C:\\reports\\images\\HFS_Logo_FullColor_RGB_Large.png')
    # Resize the image
    img.width = 200
    img.height = 108.66666
    # Add the image to the sheet
    sheet.add_image(img, 'A1')
    # Change the height of the first row
    sheet.row_dimensions[1].height = 60

    # Merge and center cells for the title
    col_count = draft_df.shape[1]
    last_column = openpyxl.utils.get_column_letter(col_count + 1)  # Adding 1 to include all columns
    sheet.merge_cells(f'B2:{last_column}2')
    sheet['B2'].alignment = Alignment(horizontal='center')

    # Set the font for the title
    font = Font(name='Calibri', size=16)
    sheet['B2'].font = font

    # Remove any borders from the title cell
    sheet['B2'].border = Border(left=Side(border_style=None),
                                right=Side(border_style=None),
                                top=Side(border_style=None),
                                bottom=Side(border_style=None))

    if sheet_name == 'Productivity':
        # Apply borders to the DataFrame cells in the 'Productivity' sheet
        df_rows, df_cols = draft_df.shape
        start_row = 6
        start_col = 2  # Column B

        # Define the end row and column
        end_row = start_row + df_rows
        end_col = start_col + df_cols - 1

        # Apply borders to the DataFrame cells
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                if col == end_col and row != start_row:
                    cell.border = cell.border + right_border
                if row == end_row:
                    cell.border = cell.border + bottom_border
                if row == start_row:
                    cell.border = Border(bottom=Side(style='thin'))
                if col == start_col and row != start_row:
                    cell.border = cell.border + left_border

        # Ensure the bottom right cell has both borders
        bottom_right_cell = sheet.cell(row=end_row, column=end_col)
        bottom_right_cell.border = bottom_border + right_border
    else:
        # Apply borders for single record sheets
        df_cols = draft_df.shape[1]
        start_row = 6
        start_col = 2  # Column B

        end_row = start_row + 1  # Single record ends one row after start_row
        end_col = start_col + df_cols - 1

        # Apply borders to the single record cells
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                if col == end_col and row != start_row:
                    cell.border = cell.border + right_border
                if row == end_row:
                    cell.border = cell.border + bottom_border
                if row == start_row:
                    cell.border = Border(bottom=Side(style='thin'))
                if col == start_col and row != start_row:
                    cell.border = cell.border + left_border

        # Ensure the bottom right cell has both borders
        bottom_right_cell = sheet.cell(row=end_row, column=end_col)
        bottom_right_cell.border = bottom_border + right_border

# Format the second sheet and the two dataframes in it
sheet = book['Scheduler_Exceptions']

# Adjust column widths and remove borders from the headers and the data for both dataframes
for df, start_col in zip([exception_block_df, scheduler_exception_legend_df], [1, num_cols_exception_block + 4]):  # Adjusted to one more column to the right
    for col in range(start_col, start_col + df.shape[1]):  # Columns start from start_col
        max_length = 0
        column = chr(64 + col)  # Convert to column letter
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2  # Add a couple of additional spaces
        sheet.column_dimensions[column].width = adjusted_width

        # Remove borders from the header and the data
        cell.border = Border(left=Side(border_style=None),
                             right=Side(border_style=None),
                             top=Side(border_style=None),
                             bottom=Side(border_style=None))

# Left align all cells in the scheduler_exception_legend_df dataframe
start_col_legend = num_cols_exception_block + 4  # Starting column for scheduler_exception_legend_df, adjusted to one more column to the right
end_col_legend = start_col_legend + scheduler_exception_legend_df.shape[1] - 1  # Ending column for scheduler_exception_legend_df
for col in range(start_col_legend, end_col_legend + 1):
    for cell in sheet[chr(64 + col)]:
        cell.alignment = Alignment(horizontal='left')

# Remove grid lines
sheet.sheet_view.showGridLines = False

# Save the changes
book.save('draft.xlsx')











from PIL import Image


# Function to resize images
def resize_image(image_path, output_path, size):
    with Image.open(image_path) as img:
        img.thumbnail(size)
        img.save(output_path, format='PNG')

# # Load the parameters from the temp_params.json file
# temp_param_file = Path('temp_params.json')
# temp_data_file = Path('temp_data.pkl')
# with open(temp_param_file, 'r') as f:
#     parameters = json.load(f)

# # Extract parameters from the loaded JSON
# to_email = parameters.get('to_email', 'default_email@example.com')
# start_date = parameters.get('start_date', '').strip()
# end_date = parameters.get('end_date', '').strip()

# # Extract email configuration from environment variables
# smtp_email = os.getenv('EMAIL_smtp_email')
# smtp_port = int(os.getenv('EMAIL_smtp_port'))
# smtp_server = os.getenv('EMAIL_smtp_server')

# Set up the SMTP server
server = smtplib.SMTP(smtp_server, smtp_port)
server.ehlo()  # Identify yourself to the SMTP server
server.starttls()
server.ehlo()  # Identify yourself to the SMTP server again after starting TLS

# Create the email
msg = MIMEMultipart('related')  # 'related' is used to send images embedded in the email
msg['From'] = smtp_email
msg['To'] = to_email

# Create the subject with the date range
if start_date and end_date:
    subject = f'Productivity Report {start_date} to {end_date}'
    body_date_info = f'for {start_date} to {end_date}'
else:
    today_date = datetime.today().strftime('%Y-%m-%d')
    subject = f'Productivity Report for {today_date}'
    body_date_info = f'for {today_date}'

msg['Subject'] = subject

# Create the body of the email
body = f"""
<html>
  <body style="font-family: 'Segoe UI', sans-serif; color: #242424;">
    <p>Hello,</p>
    <p style="margin-left: 40px;">Attached is the Productivity Report {body_date_info}.</p>
    <p>Thank you,</p>
    <div style="margin-top: 20px;">
      <table cellspacing="0" cellpadding="0" border="0" style="width: auto;">
        <tr>
          <td colspan="2">
            <hr style="border: none; border-top: 1px solid #e0e0e0; margin: 10px 0;">
          </td>
        </tr>
        <tr>
          <td style="vertical-align: top; padding-right: 10px;">
            <p style="margin: 0; font-size: 14px;"><strong>MIND</strong><br>
            Enterprise Reporting System,<br>
            Hillcrest Family Services</p>
            <p style="margin: 5px 0;">
              <a href="https://hillcrest-fs.org/" style="color: #467886; text-decoration: none;">hillcrest-fs.org</a>
            </p>
          </td>
          <td style="border-left: 1px solid #e0e0e0; padding-left: 10px; vertical-align: top;">
            <img src="cid:image2" style="height: 50px; margin-top: 10px;">
            <img src="cid:image1" style="height: 50px; margin-left: 10px; margin-top: 10px;">
          </td>
        </tr>
      </table>
    </div>
  </body>
</html>
"""
msg.attach(MIMEText(body, 'html'))

# Attach the Excel file
attachment_path = 'draft.xlsx'
attachment_name = f'Productivity Report {start_date} to {end_date}.xlsx'
part = MIMEBase('application', 'octet-stream')
with open(attachment_path, 'rb') as attachment:
    part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
msg.attach(part)

# Save a copy of the attachment to the specified directory
archive_directory = r"C:\MIND\MIND_reports\productivity_report\archived_productivity_reports"
if not os.path.exists(archive_directory):
    os.makedirs(archive_directory)
current_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
new_attachment_name = f"{os.path.splitext(attachment_name)[0]}_{current_timestamp}{os.path.splitext(attachment_name)[1]}"
shutil.copy(attachment_path, os.path.join(archive_directory, new_attachment_name))

# Resize and embed the signature image in the email
signature_image_path = 'C:\\MIND\\MIND\\MIND_images\\HFS_Logo_FullColor_RGB_Large.png'
resized_signature_image_path = 'C:\\MIND\\MIND\\MIND_images\\HFS_Logo_FullColor_RGB_Large_resized.png'
resize_image(signature_image_path, resized_signature_image_path, (200, 200))
with open(resized_signature_image_path, 'rb') as img:
    msg_image = MIMEImage(img.read())
    msg_image.add_header('Content-ID', '<image1>')
    msg.attach(msg_image)

# Resize and embed the additional image in the email
additional_image_path = 'C:\\MIND\\MIND\\MIND_images\\MIND_logo.png'
resized_additional_image_path = 'C:\\MIND\\MIND\\MIND_images\\MIND_logo_resized.png'
resize_image(additional_image_path, resized_additional_image_path, (130, 130))
with open(resized_additional_image_path, 'rb') as img:
    msg_image2 = MIMEImage(img.read())
    msg_image2.add_header('Content-ID', '<image2>')
    msg.attach(msg_image2)

# Send the email
server.send_message(msg)
server.quit()

# Delete the files
os.remove('draft.xlsx')  # Delete the draft Excel file
os.remove('output.xlsx')  # Delete the output Excel file if it's no longer needed

# Delete the resized images
os.remove(resized_signature_image_path)
os.remove(resized_additional_image_path)

print("Files deleted successfully.")
