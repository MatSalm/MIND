import pandas as pd
import configparser
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import json
from pathlib import Path

# Define the path to the config file, temp_data.pkl, and logo image
config_file_path = os.path.join('..', 'config', 'config.ini')
temp_data_path = os.path.join('.', 'temp_data.pkl')
logo_path = r'C:\MIND\MIND\MIND_images\HFS_Logo_FullColor_RGB_Large.png'

# Load the config file
config = configparser.ConfigParser()
config.read(config_file_path)

# Get the date from the config file or set it to yesterday's date
date_str = config['report'].get('yesterday_sheet', '').strip()

if not date_str:
    # If the date is not set, use yesterday's date
    yesterday_date = datetime.now() - timedelta(days=1)
    date_str = yesterday_date.strftime('%Y-%m-%d')

# Parse the date string to a datetime object
filter_date = datetime.strptime(date_str, '%Y-%m-%d')

# Load the DataFrame from the temp_data.pkl file
calendar_df = pd.read_pickle(temp_data_path)

# Ensure the 'date' column in the DataFrame is in datetime format, specifying the format
calendar_df['date'] = pd.to_datetime(calendar_df['date'], format='%Y_%m_%d')

# Filter the DataFrame for the specified date
filtered_df = calendar_df[calendar_df['date'] == filter_date]

# Define a dictionary with current column names as keys and new column names as values
column_renames = {
    'PATID': 'Client ID',
    'date': 'Date',
    'admin_hrs_default': 'Scheduled Administration Time',
    'admin_instruct_formatted': 'Administration Instructions',
    'med_descr_ext_formatted': 'Medication Dosage',
    'order_code_description': 'Medication Description',
    'program_value': 'Program'
}

# Rename the columns
filtered_df.rename(columns=column_renames, inplace=True)

from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2

# Abbreviation mapping
abbreviations = {
    'Residential Care Facility': 'RCF',
    'Adult Group Home': 'AGH',
    'Supervised Apartment Living': 'SAL',
    'Residential Care': 'RC',
    'Persistent Mental Illness': 'PMI',
}

# Function to abbreviate program names
def abbreviate_program_name(program_name):
    for key, value in abbreviations.items():
        program_name = program_name.replace(key, value)
    return program_name[:31]  # Ensure the name is no longer than 31 characters

# Create a new workbook
wb = Workbook()

# Remove the default sheet created by openpyxl
default_sheet = wb.active
wb.remove(default_sheet)

# Define border styles
thin_left_border = Border(left=Side(style='thin'))
thin_right_border = Border(right=Side(style='thin'))
thin_top_border = Border(top=Side(style='thin'))
thin_bottom_border = Border(bottom=Side(style='thin'))
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Ensure the filter_date parameter is correctly formatted
filter_date = pd.to_datetime(filter_date).strftime('%Y-%m-%d')

# Ensure the Client ID column is treated as integers
filtered_df['Client ID'] = filtered_df['Client ID'].astype(int)

# Group by Program and create sheets dynamically
program_groups = filtered_df.groupby('Program')
for program_name, group in program_groups:
    abbreviated_name = abbreviate_program_name(program_name)
    sheet = wb.create_sheet(title=abbreviated_name)
    
    # Remove gridlines
    sheet.sheet_view.showGridLines = False
    
    # Insert the logo image into cell A1
    logo = Image(logo_path)
    logo.width = int(logo.width * 0.16)
    logo.height = int(logo.height * 0.16)
    sheet.add_image(logo, 'A1')
    
    # Insert header text in merged cells E4 to I4
    header_text = f'No Documentation Med Error Report for {filter_date}'
    sheet.merge_cells('E4:I4')
    header_cell = sheet['E4']
    header_cell.value = header_text
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.font = Font(bold=True, size=16)

    # Write the DataFrame for each group to the corresponding sheet starting at cell D8
    start_row = 8
    start_col = 4
    for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='left')
            
            # Make header row bold
            if r_idx == start_row:
                cell.font = Font(bold=True)
            
            # Apply date formatting
            if cell.column_letter == 'E' and r_idx > start_row:
                cell.number_format = FORMAT_DATE_YYYYMMDD2
            
            # Apply number formatting to 'Client ID'
            if cell.column_letter == 'D' and r_idx > start_row:
                cell.number_format = '0'
    
    # Determine the max length for each column
    col_lengths = {}
    for col in range(start_col, start_col + len(group.columns)):
        max_length = 0
        for row in range(start_row, start_row + len(group) + 1):  # Include headers
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        col_lengths[col] = max_length + 2  # Add extra spaces

    # Auto-adjust column width
    for col, length in col_lengths.items():
        sheet.column_dimensions[sheet.cell(row=start_row, column=col).column_letter].width = length * 1.2

    # Add a border around the data excluding the headers
    first_row = start_row + 1
    last_row = start_row + len(group)
    first_col = start_col
    last_col = start_col + len(group.columns) - 1

    for row in range(first_row, last_row + 1):
        sheet.cell(row=row, column=first_col).border = Border(left=thin_border.left)
        sheet.cell(row=row, column=last_col).border = Border(right=thin_border.right)
        # Ensure the top border is added for the first row and bottom border for the last row
        for col in range(first_col, last_col + 1):
            if row == first_row:
                sheet.cell(row=row, column=col).border = Border(top=thin_border.top)
            if row == last_row:
                sheet.cell(row=row, column=col).border = Border(bottom=thin_border.bottom)

    # Ensure corners have the correct borders
    sheet.cell(row=first_row, column=first_col).border = Border(left=thin_border.left, top=thin_border.top)
    sheet.cell(row=first_row, column=last_col).border = Border(right=thin_border.right, top=thin_border.top)
    sheet.cell(row=last_row, column=first_col).border = Border(left=thin_border.left, bottom=thin_border.bottom)
    sheet.cell(row=last_row, column=last_col).border = Border(right=thin_border.right, bottom=thin_border.bottom)

# Save the workbook to a file with the specified filename
output_path = f'no_documentation_med_error_report_for_{filter_date}.xlsx'
wb.save(output_path)

# Save the filter_date parameter to the temp_params.json file
temp_param_file = Path('temp_params.json')
parameters = {
    'filter_date': filter_date  # Make sure filter_date is in the correct format (YYYY-MM-DD)
}

with open(temp_param_file, 'w') as f:
    json.dump(parameters, f)

print("Parameters saved successfully.")